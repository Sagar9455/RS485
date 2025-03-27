from openpyxl import Workbook
from openpyxl import load_workbook

# Create a workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Add data to the sheet
ws["A1"] = "TimeStamp"
ws["B1"] = "Description"
ws["C1"] = "Reference value"
ws["D1"] = "Actual value"
ws["E1"] = "Result"

# Open and read the space-delimited text file
with open("sampleLog.txt", "r") as file:
    lines = file.readlines()


# Process the lines and write them into the Excel sheet
for line in lines:
    # Split each line into values based on space (handling multiple spaces)
    row = line.strip().split()  # This handles multiple spaces as a delimiter
    ws.append(row)
    print(line)
    


# Save the workbook
wb.save("TestReport.xlsx")



workbook = load_workbook('TestReport.xlsx')
sheet = workbook.active

for row in sheet.iter_rows(min_row=2, max_row=5, min_col=1, max_col=5):
    if row[2].value == row[3].value:
        row[4].value = "Pass"
    else:
        row[4].value = "Fail"


workbook.save('TestReport.xlsx')
    
print("Test report generated successfully!")


