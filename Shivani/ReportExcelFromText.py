from openpyxl import Workbook

# Create a workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Add data to the sheet
ws["A1"] = "TimeStamp"
ws["B1"] = "Description"
ws["C1"] = "Reference value"
ws["D1"] = "Actual value"
ws["E1"] = "Result"



data = [
    ["02:44:54 pm 2024", "Default Session", 5001, 5001, "Pass"],
    ["02:44:55 pm 2024", "Extended Session", 5003, 5003, "Pass"],
    ["02:44:56 pm 2024", "Programming Session", 5002, 5002, "Pass"]
]

# Add rows to the sheet
for row in data:
    ws.append(row)

# Save the workbook
wb.save("openpyxl_output.xlsx")

print("Excel file generated using openpyxl!")
